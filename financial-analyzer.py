import tkinter as tk
from tkinter import filedialog, ttk, scrolledtext, messagebox
import pandas as pd
import numpy as np
import os
import json
import re
from openpyxl import load_workbook
import requests
import threading
import configparser
import tempfile
from pathlib import Path
import base64
import io
import logging

# Setup logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
#     handlers=[logging.FileHandler("financial_analyzer.log"), logging.StreamHandler()]
# )
# logger = logging.getLogger(__name__)

class GeminiModelProcessor:
    """Handles the AI processing of Excel files using Google's Gemini API"""
    
    def __init__(self, api_key=None):
        self.api_key = api_key or self._load_api_key()
        self.api_base_url = "https://generativelanguage.googleapis.com/v1beta/models"
        self.model = "gemini-2.0-flash"  # Using Gemini flash model for faster responses
    
    def _load_api_key(self):
        """Load API key from config file or environment variable"""
        config = configparser.ConfigParser()
        
        # Try to load from config file
        config_path = Path.home() / ".financial_analyzer" / "config.ini"
        if config_path.exists():
            config.read(config_path)
            if "API" in config and "gemini_api_key" in config["API"]:
                return config["API"]["gemini_api_key"]
        
        # Try to load from environment variable
        import os
        return os.environ.get("GEMINI_API_KEY")
    
    def save_api_key(self, api_key):
        """Save API key to config file"""
        config = configparser.ConfigParser()
        
        # Create directory if it doesn't exist
        config_dir = Path.home() / ".financial_analyzer"
        config_dir.mkdir(exist_ok=True, parents=True)
        
        config_path = config_dir / "config.ini"
        
        # Load existing config if it exists
        if config_path.exists():
            config.read(config_path)
        
        # Ensure API section exists
        if "API" not in config:
            config["API"] = {}
        
        # Set API key
        config["API"]["gemini_api_key"] = api_key
        
        # Write to file
        with open(config_path, "w") as f:
            config.write(f)
        
        self.api_key = api_key
    
    def _extract_and_prepare_data(self, file_path):
        """Extract key information from Excel file and prepare it for analysis"""
        try:
            # Load workbook
            wb = load_workbook(file_path, read_only=True, data_only=True)
            
            excel_structure = {
                "filename": os.path.basename(file_path),
                "sheets": []
            }
            
            # Process each sheet
            for sheet_name in wb.sheetnames:
                sheet = wb[sheet_name]
                
                # Extract sheet data (limited sample)
                data = []
                row_count = 0
                for row in sheet.iter_rows(values_only=True):
                    if row_count > 100:  # Limit to 100 rows per sheet for efficiency
                        break
                    # Convert any non-serializable types to strings
                    processed_row = []
                    for cell in row:
                        if cell is None:
                            processed_row.append(None)
                        else:
                            try:
                                # Test if JSON serializable
                                json.dumps(cell)
                                processed_row.append(cell)
                            except (TypeError, OverflowError):
                                # Convert to string if not serializable
                                processed_row.append(str(cell))
                    data.append(processed_row)
                    row_count += 1
                
                # Add sheet info
                excel_structure["sheets"].append({
                    "name": sheet_name,
                    "data": data
                })
            
            # Create structured data for AI
            return excel_structure
            
        except Exception as e:
            logger.error(f"Error extracting data from Excel: {e}", exc_info=True)
            raise Exception(f"Failed to process Excel file: {str(e)}")
    
    def analyze_excel_file(self, file_path, progress_callback=None):
        """
        Analyze Excel file using Gemini AI to extract financial model information
        
        Args:
            file_path: Path to Excel file
            progress_callback: Callback function to update progress
        
        Returns:
            Dictionary containing analysis results
        """
        if progress_callback:
            progress_callback("Preparing Excel file for analysis...")
        
        # Check if API key is set
        if not self.api_key:
            raise ValueError("Gemini API key is not set. Please set it in Settings.")
        
        try:
            # Extract data from Excel file
            excel_data = self._extract_and_prepare_data(file_path)
            
            if progress_callback:
                progress_callback("Sending data to Gemini AI for analysis...")
            
            # Analyze with AI
            return self._analyze_with_gemini(excel_data, progress_callback)
                
        except Exception as e:
            logger.error(f"Error analyzing Excel file: {e}", exc_info=True)
            raise
    
    def _analyze_with_gemini(self, excel_data, progress_callback=None):
        """Analyze Excel data using Gemini API"""
        if progress_callback:
            progress_callback("Processing with Gemini AI model...")
        
        # Convert data to JSON
        excel_data_json = json.dumps(excel_data)
        
        system_prompt = """
            You are a financial model analysis expert. Your task is to analyze the Excel structure data provided
            and extract the following information:
            
            1. Key assumptions used in the model (look for inputs, parameters, rates, growth values, etc.)
            2. Financial returns (NPV, IRR, ROI, payback period, profit margins, etc.)
            3. Cash flow projections and summary
            
            The data provided will include sheet names and rows from each sheet.
            Use this information to identify financial model components and extract meaningful information.
            
            Look for patterns in the data that indicate:
            - Input parameters or assumptions (often in dedicated sheets or sections)
            - Calculation results showing financial returns
            - Time series data showing cash flows or projections
            - Summary metrics and KPIs
            
            Provide your analysis in a structured JSON format as follows:
            {
                "assumptions": [
                    {"description": "Description of assumption", "value": "Value of assumption"}
                ],
                "financial_returns": {
                    "npv": {"label": "NPV label as found", "value": "NPV value"},
                    "irr": {"label": "IRR label as found", "value": "IRR value"},
                    "payback_period": {"label": "Payback period label as found", "value": "Payback period value"},
                    "roi": {"label": "ROI label as found", "value": "ROI value"},
                    "profit_margin": {"label": "Profit margin label as found", "value": "Profit margin value"},
                    "other_metrics": [
                        {"label": "Other metric label", "value": "Other metric value"}
                    ]
                },
                "cash_flows": [
                    {
                        "label": "Cash flow label",
                        "periods": [
                            {"period": "Period identifier", "value": "Cash flow value"}
                        ]
                    }
                ],
                "summary": "A text summary of the financial model analysis, including your interpretation of the model's purpose, key metrics, and overall financial outlook based on the data."
            }
            
            Be thorough in examining all sheets and their data. If you can't find specific information, indicate that in your response.
            Ensure your response is valid JSON and only includes the JSON. Do not include any other text before or after the JSON.
            """
        
        user_prompt = "Please analyze this Excel financial model data and extract key information about assumptions, financial returns, and cash flows. Provide the analysis in the specified JSON format."
        
        # Add truncated data hint to help the AI
        if len(excel_data["sheets"]) > 0:
            sample_count = sum(1 for sheet in excel_data["sheets"] if len(sheet.get("data", [])) >= 100)
            if sample_count > 0:
                user_prompt += f" Note that {sample_count} sheet(s) were truncated to the first 100 rows to manage data size."
        
        # Prepare request for Gemini API
        url = f"{self.api_base_url}/{self.model}:generateContent?key={self.api_key}"
        
        # Gemini API expects a different format than OpenAI
        data = {
            "contents": [
                {
                    "parts": [
                        {"text": system_prompt},
                        {"text": f"{user_prompt}\n\nExcel Data: {excel_data_json}"}
                    ]
                }
            ],
            "generationConfig": {
                "temperature": 0.2,
                "topP": 0.8,
                "topK": 40,
                "maxOutputTokens": 8192,
                "responseMimeType": "application/json"
            }
        }
        
        headers = {
            "Content-Type": "application/json"
        }
        
        try:
            response = requests.post(url, headers=headers, json=data)
            
            if response.status_code != 200:
                logger.error(f"Gemini API error: {response.status_code} - {response.text}")
                raise Exception(f"Gemini API error: {response.status_code} - {response.text}")
            
            response_data = response.json()
            
            # Extract content from Gemini response
            if 'candidates' in response_data and len(response_data['candidates']) > 0:
                candidate = response_data['candidates'][0]
                if 'content' in candidate and 'parts' in candidate['content']:
                    parts = candidate['content']['parts']
                    if len(parts) > 0 and 'text' in parts[0]:
                        ai_response = parts[0]['text']
                    else:
                        raise Exception("Unexpected response format from Gemini API")
                else:
                    raise Exception("Unexpected response format from Gemini API")
            else:
                raise Exception("No content returned from Gemini API")
            
            # Process the response to extract JSON
            try:
                # Try to find JSON in the response
                json_start = ai_response.find('{')
                json_end = ai_response.rfind('}') + 1
                
                if json_start != -1 and json_end != -1:
                    json_str = ai_response[json_start:json_end]
                    analysis_results = json.loads(json_str)
                else:
                    # If no JSON brackets found, try the whole response
                    analysis_results = json.loads(ai_response)
                
                return analysis_results
            except json.JSONDecodeError as e:
                logger.error(f"Error parsing JSON response: {e}", exc_info=True)
                logger.error(f"Raw response: {ai_response[:500]}...")  # Log part of the response for debugging
                raise Exception(f"Error parsing Gemini response as JSON: {e}")
                
        except requests.exceptions.RequestException as e:
            logger.error(f"Request error: {e}", exc_info=True)
            raise Exception(f"Failed to communicate with Gemini API: {str(e)}")


class FinancialModelAnalyzer:
    def __init__(self, root):
        self.root = root
        self.root.title("Gemini AI Financial Model Analyzer")
        self.root.geometry("900x700")
        self.root.minsize(900, 700)
        
        # Create AI processor
        self.ai_processor = GeminiModelProcessor()
        
        # Set up the main frame
        self.main_frame = ttk.Frame(root, padding="20")
        self.main_frame.pack(fill=tk.BOTH, expand=True)
        
        # Application title
        title_label = ttk.Label(self.main_frame, text="Gemini AI Financial Model Analyzer", font=("Arial", 16, "bold"))
        title_label.pack(pady=(0, 10))
        
        # Description
        desc_label = ttk.Label(self.main_frame, 
                               text="Upload your Excel financial model to extract key assumptions and financial returns using Gemini AI",
                               wraplength=800)
        desc_label.pack(pady=(0, 20))
        
        # File selection frame
        self.file_frame = ttk.Frame(self.main_frame)
        self.file_frame.pack(fill=tk.X, pady=(0, 20))
        
        # File path entry
        self.file_path_var = tk.StringVar()
        self.file_path_entry = ttk.Entry(self.file_frame, textvariable=self.file_path_var, width=60)
        self.file_path_entry.pack(side=tk.LEFT, padx=(0, 10), fill=tk.X, expand=True)
        
        # Browse button
        self.browse_button = ttk.Button(self.file_frame, text="Browse", command=self.browse_file)
        self.browse_button.pack(side=tk.LEFT, padx=(0, 10))
        
        # Analyze button
        self.analyze_button = ttk.Button(self.file_frame, text="Analyze with Gemini", command=self.analyze_file)
        self.analyze_button.pack(side=tk.LEFT)
        
        # Settings button
        self.settings_button = ttk.Button(self.file_frame, text="Settings", command=self.open_settings)
        self.settings_button.pack(side=tk.LEFT, padx=(10, 0))
        
        # Progress bar
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(self.main_frame, variable=self.progress_var, mode='indeterminate')
        self.progress_bar.pack(fill=tk.X, pady=(0, 10))
        
        # Create a notebook for tabs
        self.notebook = ttk.Notebook(self.main_frame)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        # Create tabs
        self.summary_tab = ttk.Frame(self.notebook)
        self.assumptions_tab = ttk.Frame(self.notebook)
        self.returns_tab = ttk.Frame(self.notebook)
        self.cashflow_tab = ttk.Frame(self.notebook)
        
        self.notebook.add(self.summary_tab, text="Summary")
        self.notebook.add(self.assumptions_tab, text="Assumptions")
        self.notebook.add(self.returns_tab, text="Financial Returns")
        self.notebook.add(self.cashflow_tab, text="Cash Flows")
        
        # Summary Text Area
        self.summary_text = scrolledtext.ScrolledText(self.summary_tab, wrap=tk.WORD)
        self.summary_text.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create treeviews for other tabs
        self.setup_assumptions_tab()
        self.setup_returns_tab()
        self.setup_cashflows_tab()
        
        # Status bar
        self.status_var = tk.StringVar()
        self.status_var.set("Ready. Please select an Excel file.")
        self.status_bar = ttk.Label(self.root, textvariable=self.status_var, relief=tk.SUNKEN, anchor=tk.W)
        self.status_bar.pack(side=tk.BOTTOM, fill=tk.X)
        
        # Initialize results dictionary
        self.results = {
            "assumptions": [],
            "financial_returns": {
                "npv": None,
                "irr": None,
                "payback_period": None,
                "roi": None,
                "profit_margin": None,
                "other_metrics": []
            },
            "cash_flows": [],
            "summary": ""
        }
    
    def setup_assumptions_tab(self):
        # Create frame for treeview
        frame = ttk.Frame(self.assumptions_tab)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Create treeview
        columns = ("description", "value")
        self.assumptions_tree = ttk.Treeview(frame, columns=columns, show="headings", yscrollcommand=scrollbar.set)
        
        # Configure scrollbar
        scrollbar.config(command=self.assumptions_tree.yview)
        
        # Set column headings
        self.assumptions_tree.heading("description", text="Description")
        self.assumptions_tree.heading("value", text="Value")
        
        # Set column widths
        self.assumptions_tree.column("description", width=400)
        self.assumptions_tree.column("value", width=200)
        
        # Pack treeview
        self.assumptions_tree.pack(fill=tk.BOTH, expand=True)
    
    def setup_returns_tab(self):
        # Create frame for treeview
        frame = ttk.Frame(self.returns_tab)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Create treeview
        columns = ("metric", "value")
        self.returns_tree = ttk.Treeview(frame, columns=columns, show="headings", yscrollcommand=scrollbar.set)
        
        # Configure scrollbar
        scrollbar.config(command=self.returns_tree.yview)
        
        # Set column headings
        self.returns_tree.heading("metric", text="Metric")
        self.returns_tree.heading("value", text="Value")
        
        # Set column widths
        self.returns_tree.column("metric", width=400)
        self.returns_tree.column("value", width=200)
        
        # Pack treeview
        self.returns_tree.pack(fill=tk.BOTH, expand=True)
    
    def setup_cashflows_tab(self):
        # Create frame for treeview
        frame = ttk.Frame(self.cashflow_tab)
        frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        # Create scrollbar
        scrollbar = ttk.Scrollbar(frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        # Create treeview
        columns = ("description", "start_value", "end_value")
        self.cashflow_tree = ttk.Treeview(frame, columns=columns, show="headings", yscrollcommand=scrollbar.set)
        
        # Configure scrollbar
        scrollbar.config(command=self.cashflow_tree.yview)
        
        # Set column headings
        self.cashflow_tree.heading("description", text="Description")
        self.cashflow_tree.heading("start_value", text="Starting Value")
        self.cashflow_tree.heading("end_value", text="Ending Value")
        
        # Set column widths
        self.cashflow_tree.column("description", width=300)
        self.cashflow_tree.column("start_value", width=150)
        self.cashflow_tree.column("end_value", width=150)
        
        # Pack treeview
        self.cashflow_tree.pack(fill=tk.BOTH, expand=True)
    
    def browse_file(self):
        file_path = filedialog.askopenfilename(
            filetypes=[("Excel files", "*.xlsx *.xls")]
        )
        if file_path:
            self.file_path_var.set(file_path)
            self.status_var.set(f"Selected file: {os.path.basename(file_path)}")
    
    def analyze_file(self):
        file_path = self.file_path_var.get()
        
        if not file_path:
            self.status_var.set("Error: Please select a file first")
            return
        
        try:
            # Start progress bar
            self.progress_bar.start(10)
            self.analyze_button.configure(state="disabled")
            self.browse_button.configure(state="disabled")
            
            # Reset previous results
            self.clear_results()
            
            # Use threading to prevent UI freezing during API call
            self.analysis_thread = threading.Thread(
                target=self._run_analysis, 
                args=(file_path,)
            )
            self.analysis_thread.daemon = True
            self.analysis_thread.start()
            
        except Exception as e:
            self.status_var.set(f"Error: {str(e)}")
            
            # Stop progress bar
            self.progress_bar.stop()
            self.analyze_button.configure(state="normal")
            self.browse_button.configure(state="normal")
            
            logger.error(f"Error analyzing file: {e}", exc_info=True)
    
    def _run_analysis(self, file_path):
        try:
            # Update status
            self.root.after(0, lambda: self.status_var.set("Analyzing Excel file with Gemini AI..."))
            
            # Analyze with AI
            results = self.ai_processor.analyze_excel_file(
                file_path, 
                progress_callback=lambda msg: self.root.after(0, lambda: self.status_var.set(msg))
            )
            
            # Update UI with results
            self.root.after(0, lambda: self._update_ui_with_results(results))
            
            # Update status
            self.root.after(0, lambda: self.status_var.set("Analysis complete"))
            
        except Exception as e:
            self.root.after(0, lambda: self.status_var.set(f"Error: {str(e)}"))
            logger.error(f"Error in analysis thread: {e}", exc_info=True)
        finally:
            # Stop progress bar
            self.root.after(0, lambda: self.progress_bar.stop())
            self.root.after(0, lambda: self.analyze_button.configure(state="normal"))
            self.root.after(0, lambda: self.browse_button.configure(state="normal"))
    
    def _update_ui_with_results(self, results):
        # Update the results dictionary
        self.results = results
        
        # Display results
        self.display_results()
    
    def clear_results(self):
        # Clear results dictionary
        self.results = {
            "assumptions": [],
            "financial_returns": {
                "npv": None,
                "irr": None,
                "payback_period": None,
                "roi": None,
                "profit_margin": None,
                "other_metrics": []
            },
            "cash_flows": [],
            "summary": ""
        }
        
        # Clear UI elements
        self.summary_text.delete(1.0, tk.END)
        
        # Clear treeviews
        for item in self.assumptions_tree.get_children():
            self.assumptions_tree.delete(item)
        
        for item in self.returns_tree.get_children():
            self.returns_tree.delete(item)
        
        for item in self.cashflow_tree.get_children():
            self.cashflow_tree.delete(item)
    
    def display_results(self):
        # Display summary
        if "summary" in self.results and self.results["summary"]:
            self.summary_text.insert(tk.END, self.results["summary"])
        else:
            self.summary_text.insert(tk.END, "No summary generated by AI analysis.")
        
        # Display assumptions
        if "assumptions" in self.results:
            for assumption in self.results["assumptions"]:
                if "description" in assumption and "value" in assumption:
                    self.assumptions_tree.insert("", tk.END, values=(
                        assumption["description"], 
                        assumption["value"]
                    ))
        
        # Display financial returns
        if "financial_returns" in self.results:
            for key, value in self.results["financial_returns"].items():
                if key == "other_metrics" and isinstance(value, list):
                    for metric in value:
                        if "label" in metric and "value" in metric:
                            self.returns_tree.insert("", tk.END, values=(
                                metric["label"], 
                                metric["value"]
                            ))
                elif value is not None and isinstance(value, dict) and "label" in value and "value" in value:
                    self.returns_tree.insert("", tk.END, values=(
                        value["label"], 
                        value["value"]
                    ))
        
        # Display cash flows
        if "cash_flows" in self.results:
            for cf in self.results["cash_flows"]:
                if "label" in cf and "periods" in cf and cf["periods"]:
                    periods = cf["periods"]
                    first = periods[0] if len(periods) > 0 else {"value": "N/A"}
                    last = periods[-1] if len(periods) > 0 else {"value": "N/A"}
                    
                    self.cashflow_tree.insert("", tk.END, values=(
                        cf["label"],
                        first.get("value", "N/A"),
                        last.get("value", "N/A")
                    ))
    
    def open_settings(self):
        """Open settings dialog"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("Settings")
        settings_window.geometry("400x250")  # Made taller to ensure enough space
        settings_window.minsize(400, 250)
        settings_window.transient(self.root)
        settings_window.grab_set()
        
        # Create frame
        frame = ttk.Frame(settings_window, padding="20")
        frame.pack(fill=tk.BOTH, expand=True)
        
        # API Key Entry
        ttk.Label(frame, text="Gemini API Key:").pack(anchor=tk.W, pady=(0, 5))
        
        api_key_var = tk.StringVar()
        api_key_var.set(self.ai_processor.api_key or "")
        
        api_key_entry = ttk.Entry(frame, textvariable=api_key_var, width=40, show="*")
        api_key_entry.pack(fill=tk.X, pady=(0, 15))
        
        # Show/Hide API Key
        show_api_key_var = tk.BooleanVar()
        show_api_key_var.set(False)
        
        show_api_key_cb = ttk.Checkbutton(
            frame, 
            text="Show API Key", 
            variable=show_api_key_var,
            command=lambda: api_key_entry.configure(show="" if show_api_key_var.get() else "*")
        )
        show_api_key_cb.pack(anchor=tk.W, pady=(0, 15))
        
        # API help text
        help_text = ttk.Label(
            frame, 
            text="Get a Gemini API key from Google AI Studio\nhttps://makersuite.google.com/app/apikey",
            wraplength=350, 
            justify=tk.LEFT, 
            font=("Arial", 8)
        )
        help_text.pack(anchor=tk.W, pady=(0, 15))
        
        # Button frame to ensure visibility
        button_frame = ttk.Frame(frame)
        button_frame.pack(fill=tk.X, pady=(5, 0))
        
        # Save button with better styling
        save_button = ttk.Button(
            button_frame, 
            text="Save", 
            command=lambda: self._save_settings(api_key_var.get(), settings_window)
        )
        save_button.pack(padx=5, pady=5, ipadx=10, ipady=5)  # Added padding for better visibility
    
    def _save_settings(self, api_key, settings_window):
        """Save settings and close dialog"""
        try:
            self.ai_processor.save_api_key(api_key)
            self.status_var.set("Settings saved successfully")
            settings_window.destroy()
        except Exception as e:
            messagebox.showerror("Error", f"Error saving settings: {str(e)}")
            logger.error(f"Error saving settings: {e}", exc_info=True)


def main():
    root = tk.Tk()
    app = FinancialModelAnalyzer(root)
    root.mainloop()


if __name__ == "__main__":
    main()