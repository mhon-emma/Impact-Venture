import pandas as pd
from openpyxl import load_workbook
import re

# Load the Excel file
file_path = "Financial_model.xlsx"
wb = load_workbook(filename=file_path, data_only=True)  # data_only=True to get values
wb_formulas = load_workbook(filename=file_path, data_only=False)  # To get formulas

# Lists to store our keyword-number pairs
hardcoded_pairs = []
formula_pairs = []

# Function to check if a string could be a keyword
def is_potential_keyword(text):
    if not isinstance(text, str):
        return False
    # Not a formula, not just numeric
    return not text.startswith('=') and not re.match(r'^-?\d+(\.\d+)?$', text.strip())

# Scan all sheets for keyword-number pairs
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    ws_formulas = wb_formulas[sheet_name]
    
    # Get all rows as lists
    rows = list(ws.values)
    
    # Scan for adjacent keyword-number pairs in rows
    for i, row in enumerate(rows):
        if not row:
            continue
            
        for j in range(len(row) - 1):
            # Check if we have a potential keyword followed by a number
            if (is_potential_keyword(row[j]) and 
                isinstance(row[j+1], (int, float))):
                
                keyword = row[j]
                number = row[j+1]
                
                # Get the cell coordinates
                cell_coord = f"{chr(65 + j)}{i + 1}"  # Convert to A1 notation
                num_cell_coord = f"{chr(65 + j + 1)}{i + 1}"
                
                # Check if the number cell contains a formula in the formula workbook
                formula_cell = ws_formulas[num_cell_coord]
                
                # In newer versions of openpyxl, we need to access the cell.value attribute when it's a formula
                if formula_cell.value and isinstance(formula_cell.value, str) and formula_cell.value.startswith('='):
                    # This is a formula-generated number
                    formula_pairs.append({
                        "keyword": keyword,
                        "number": number,
                        "formula": formula_cell.value,  # Use value instead of formula
                        "location": f"{sheet_name}!{num_cell_coord}"
                    })
                else:
                    # This is a hardcoded number
                    hardcoded_pairs.append({
                        "keyword": keyword,
                        "number": number,
                        "location": f"{sheet_name}!{num_cell_coord}"
                    })

# Print the results
print("HARDCODED KEYWORD-NUMBER PAIRS:")
for pair in hardcoded_pairs:
    print(f"{pair['keyword']}: {pair['number']} (at {pair['location']})")

print("\nFORMULA-GENERATED KEYWORD-NUMBER PAIRS:")
for pair in formula_pairs:
    print(f"{pair['keyword']}: {pair['number']} (formula: {pair['formula']} at {pair['location']})")

# Create two lists in the format you requested
# 1. Hardcoded list with keyword-number pairs
hardcoded_list = {pair["keyword"]: pair["number"] for pair in hardcoded_pairs}

# 2. Formula-generated list with keyword-number pairs
formula_list = {pair["keyword"]: pair["number"] for pair in formula_pairs}

print("\nHardcoded list:")
print(hardcoded_list)

print("\nFormula-generated list:")
print(formula_list)