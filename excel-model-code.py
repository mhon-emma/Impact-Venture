import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
import os

# Create a sample financial model Excel file
def create_financial_model(filename='sample_financial_model.xlsx'):
    # Create a workbook and sheets
    workbook = openpyxl.Workbook()
    
    # Rename the default sheet to "Assumptions"
    assumptions_sheet = workbook.active
    assumptions_sheet.title = "Assumptions"
    
    # Create additional sheets
    projections_sheet = workbook.create_sheet(title="Projections")
    returns_sheet = workbook.create_sheet(title="Financial Returns")
    
    # Define some styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    center_alignment = Alignment(horizontal='center')
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # ----- ASSUMPTIONS SHEET -----
    # Add title
    assumptions_sheet['A1'] = "PROJECT FINANCIAL MODEL - ASSUMPTIONS"
    assumptions_sheet['A1'].font = Font(bold=True, size=14)
    assumptions_sheet.merge_cells('A1:D1')
    
    # General assumptions section
    assumptions_sheet['A3'] = "General Assumptions"
    assumptions_sheet['A3'].font = Font(bold=True, size=12)
    
    # Add assumption rows
    assumptions = [
        ("Project Timeline (Years)", 5),
        ("Discount Rate", "12.0%"),
        ("Tax Rate", "25.0%"),
        ("Inflation Rate", "2.5%"),
        ("Initial Investment", "$1,000,000"),
    ]
    
    for idx, (label, value) in enumerate(assumptions, start=4):
        assumptions_sheet[f'A{idx}'] = label
        assumptions_sheet[f'B{idx}'] = value
    
    # Revenue assumptions section
    assumptions_sheet['A10'] = "Revenue Assumptions"
    assumptions_sheet['A10'].font = Font(bold=True, size=12)
    
    revenue_assumptions = [
        ("Units Sold - Year 1", 10000),
        ("Unit Price", "$50"),
        ("Unit Growth Rate", "15.0%"),
        ("Price Growth Rate", "3.0%"),
    ]
    
    for idx, (label, value) in enumerate(revenue_assumptions, start=11):
        assumptions_sheet[f'A{idx}'] = label
        assumptions_sheet[f'B{idx}'] = value
    
    # Cost assumptions section
    assumptions_sheet['A16'] = "Cost Assumptions"
    assumptions_sheet['A16'].font = Font(bold=True, size=12)
    
    cost_assumptions = [
        ("Variable Cost per Unit", "$20"),
        ("Fixed Costs - Year 1", "$200,000"),
        ("Fixed Cost Growth Rate", "5.0%"),
    ]
    
    for idx, (label, value) in enumerate(cost_assumptions, start=17):
        assumptions_sheet[f'A{idx}'] = label
        assumptions_sheet[f'B{idx}'] = value
    
    # Format columns
    for col in ['A', 'B']:
        for row in range(3, 20):
            cell = assumptions_sheet[f'{col}{row}']
            cell.border = border
    
    # ----- PROJECTIONS SHEET -----
    # Add title
    projections_sheet['A1'] = "PROJECT FINANCIAL PROJECTIONS"
    projections_sheet['A1'].font = Font(bold=True, size=14)
    projections_sheet.merge_cells('A1:G1')
    
    # Add year headers
    projections_sheet['A3'] = "Year"
    for year in range(6):  # 0 to 5 (initial + 5 years)
        projections_sheet[f'{get_column_letter(year+2)}3'] = f"Year {year}"
    
    # Format headers
    for col in range(1, 8):
        cell = projections_sheet[f'{get_column_letter(col)}3']
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
    
    # Calculate projections
    initial_units = 10000
    unit_growth = 0.15
    initial_price = 50
    price_growth = 0.03
    variable_cost = 20
    initial_fixed_cost = 200000
    fixed_cost_growth = 0.05
    
    # Units projection
    projections_sheet['A4'] = "Units Sold"
    for year in range(6):
        if year == 0:
            units = 0  # No units in initial year
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
        projections_sheet[f'{get_column_letter(year+2)}4'] = int(units)
    
    # Price projection
    projections_sheet['A5'] = "Unit Price"
    for year in range(6):
        if year == 0:
            price = 0  # No price in initial year
        else:
            price = initial_price * (1 + price_growth) ** (year - 1)
        projections_sheet[f'{get_column_letter(year+2)}5'] = f"${price:.2f}"
    
    # Revenue projection
    projections_sheet['A6'] = "Revenue"
    for year in range(6):
        if year == 0:
            revenue = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
        projections_sheet[f'{get_column_letter(year+2)}6'] = f"${revenue:,.2f}"
    
    # Variable Costs
    projections_sheet['A8'] = "Variable Costs"
    for year in range(6):
        if year == 0:
            var_costs = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            var_costs = units * variable_cost
        projections_sheet[f'{get_column_letter(year+2)}8'] = f"${var_costs:,.2f}"
    
    # Fixed Costs
    projections_sheet['A9'] = "Fixed Costs"
    for year in range(6):
        if year == 0:
            fixed_costs = 0
        else:
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
        projections_sheet[f'{get_column_letter(year+2)}9'] = f"${fixed_costs:,.2f}"
    
    # Total Costs
    projections_sheet['A10'] = "Total Costs"
    for year in range(6):
        if year == 0:
            total_costs = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
        projections_sheet[f'{get_column_letter(year+2)}10'] = f"${total_costs:,.2f}"
    
    # Operating Income
    projections_sheet['A12'] = "Operating Income"
    for year in range(6):
        if year == 0:
            op_income = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
            op_income = revenue - total_costs
        projections_sheet[f'{get_column_letter(year+2)}12'] = f"${op_income:,.2f}"
    
    # Taxes
    projections_sheet['A13'] = "Taxes (25%)"
    for year in range(6):
        if year == 0:
            taxes = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
            op_income = revenue - total_costs
            taxes = max(0, op_income * 0.25)  # Only tax positive income
        projections_sheet[f'{get_column_letter(year+2)}13'] = f"${taxes:,.2f}"
    
    # Net Income
    projections_sheet['A14'] = "Net Income"
    for year in range(6):
        if year == 0:
            net_income = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
            op_income = revenue - total_costs
            taxes = max(0, op_income * 0.25)
            net_income = op_income - taxes
        projections_sheet[f'{get_column_letter(year+2)}14'] = f"${net_income:,.2f}"
    
    # Cash Flow Section
    projections_sheet['A16'] = "CASH FLOW ANALYSIS"
    projections_sheet['A16'].font = Font(bold=True)
    
    # Initial Investment
    projections_sheet['A17'] = "Initial Investment"
    projections_sheet[f'B17'] = "-$1,000,000"
    for year in range(1, 6):
        projections_sheet[f'{get_column_letter(year+2)}17'] = "$0"
    
    # Operating Cash Flow (same as Net Income for simplicity)
    projections_sheet['A18'] = "Operating Cash Flow"
    for year in range(6):
        if year == 0:
            op_cash = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
            op_income = revenue - total_costs
            taxes = max(0, op_income * 0.25)
            op_cash = op_income - taxes
        projections_sheet[f'{get_column_letter(year+2)}18'] = f"${op_cash:,.2f}"
    
    # Net Cash Flow
    projections_sheet['A19'] = "Net Cash Flow"
    for year in range(6):
        if year == 0:
            if year == 0:
                net_cash = -1000000
            else:
                net_cash = 0
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
            op_income = revenue - total_costs
            taxes = max(0, op_income * 0.25)
            op_cash = op_income - taxes
            net_cash = op_cash
        projections_sheet[f'{get_column_letter(year+2)}19'] = f"${net_cash:,.2f}"
    
    # Cumulative Cash Flow
    projections_sheet['A20'] = "Cumulative Cash Flow"
    cumulative = 0
    for year in range(6):
        if year == 0:
            cumulative = -1000000
        else:
            units = initial_units * (1 + unit_growth) ** (year - 1)
            price = initial_price * (1 + price_growth) ** (year - 1)
            revenue = units * price
            var_costs = units * variable_cost
            fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
            total_costs = var_costs + fixed_costs
            op_income = revenue - total_costs
            taxes = max(0, op_income * 0.25)
            op_cash = op_income - taxes
            cumulative += op_cash
        projections_sheet[f'{get_column_letter(year+2)}20'] = f"${cumulative:,.2f}"
    
    # Format projections cells
    for row in range(4, 21):
        for col in range(1, 8):
            cell = projections_sheet[f'{get_column_letter(col)}{row}']
            cell.border = border
    
    # ----- FINANCIAL RETURNS SHEET -----
    # Add title
    returns_sheet['A1'] = "FINANCIAL RETURNS ANALYSIS"
    returns_sheet['A1'].font = Font(bold=True, size=14)
    returns_sheet.merge_cells('A1:C1')
    
    # Calculate NPV
    npv = -1000000  # Initial investment
    discount_rate = 0.12
    
    for year in range(1, 6):
        units = initial_units * (1 + unit_growth) ** (year - 1)
        price = initial_price * (1 + price_growth) ** (year - 1)
        revenue = units * price
        var_costs = units * variable_cost
        fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
        total_costs = var_costs + fixed_costs
        op_income = revenue - total_costs
        taxes = max(0, op_income * 0.25)
        op_cash = op_income - taxes
        
        npv += op_cash / ((1 + discount_rate) ** year)
    
    # IRR calculation (simplified approach)
    # For a proper IRR, you'd use numpy's IRR function, but for simplicity we'll estimate
    # We'll just calculate a rough ROI
    total_cash_inflow = 0
    for year in range(1, 6):
        units = initial_units * (1 + unit_growth) ** (year - 1)
        price = initial_price * (1 + price_growth) ** (year - 1)
        revenue = units * price
        var_costs = units * variable_cost
        fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
        total_costs = var_costs + fixed_costs
        op_income = revenue - total_costs
        taxes = max(0, op_income * 0.25)
        op_cash = op_income - taxes
        total_cash_inflow += op_cash
    
    irr = (total_cash_inflow / 1000000) ** (1/5) - 1  # Simplified IRR calculation
    
    # Payback period (simplified calculation)
    cumulative = -1000000
    payback_year = 0
    for year in range(1, 6):
        units = initial_units * (1 + unit_growth) ** (year - 1)
        price = initial_price * (1 + price_growth) ** (year - 1)
        revenue = units * price
        var_costs = units * variable_cost
        fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
        total_costs = var_costs + fixed_costs
        op_income = revenue - total_costs
        taxes = max(0, op_income * 0.25)
        op_cash = op_income - taxes
        
        cumulative += op_cash
        if cumulative >= 0 and payback_year == 0:
            payback_year = year
    
    # Add financial metrics
    returns_sheet['A3'] = "Financial Metric"
    returns_sheet['B3'] = "Value"
    returns_sheet['C3'] = "Notes"
    
    returns_sheet['A4'] = "Net Present Value (NPV)"
    returns_sheet['B4'] = f"${npv:,.2f}"
    returns_sheet['C4'] = "Discount Rate: 12%"
    
    returns_sheet['A5'] = "Internal Rate of Return (IRR)"
    returns_sheet['B5'] = f"{irr:.2%}"
    returns_sheet['C5'] = "Annualized return"
    
    returns_sheet['A6'] = "Payback Period"
    returns_sheet['B6'] = f"{payback_year} years"
    returns_sheet['C6'] = "Years to recover initial investment"
    
    returns_sheet['A7'] = "Return on Investment (ROI)"
    roi = (total_cash_inflow - 1000000) / 1000000
    returns_sheet['B7'] = f"{roi:.2%}"
    returns_sheet['C7'] = "Total return over 5 years"
    
    returns_sheet['A8'] = "Profit Margin"
    # Calculate average profit margin across years
    total_revenue = 0
    total_net_income = 0
    for year in range(1, 6):
        units = initial_units * (1 + unit_growth) ** (year - 1)
        price = initial_price * (1 + price_growth) ** (year - 1)
        revenue = units * price
        var_costs = units * variable_cost
        fixed_costs = initial_fixed_cost * (1 + fixed_cost_growth) ** (year - 1)
        total_costs = var_costs + fixed_costs
        op_income = revenue - total_costs
        taxes = max(0, op_income * 0.25)
        net_income = op_income - taxes
        
        total_revenue += revenue
        total_net_income += net_income
    
    avg_profit_margin = total_net_income / total_revenue
    returns_sheet['B8'] = f"{avg_profit_margin:.2%}"
    returns_sheet['C8'] = "Average over 5 years"
    
    # Format financial returns cells
    for row in range(3, 9):
        for col in range(1, 4):
            cell = returns_sheet[f'{get_column_letter(col)}{row}']
            cell.border = border
    
    # Format header row
    for col in range(1, 4):
        cell = returns_sheet[f'{get_column_letter(col)}3']
        cell.font = header_font
        cell.fill = header_fill
    
    # Set column widths
    for sheet in [assumptions_sheet, projections_sheet, returns_sheet]:
        dim_holder = DimensionHolder(worksheet=sheet)
        
        for col in range(sheet.min_column, sheet.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(sheet, min=col, max=col, width=20)
            
        sheet.column_dimensions = dim_holder
    
    # Save the workbook
    workbook.save(filename)
    return filename

# Create the sample financial model
if __name__ == "__main__":
    filename = create_financial_model('sample_financial_model.xlsx')
    print(f"Created sample financial model: {filename}")
    print(f"Current working directory: {os.getcwd()}")
